import openpyxl
import os


def write_row_name(row_names, excel_name='compare_goals_actions.xlsx', sheet_name='All Example'):
    """
    按列表输入行名称
    input：  row_names, excel_name='compare_goals_actions.xlsx', sheet_name='All Example'

    """
    # 查找同名excel,若无创建一个
    if not os.path.exists(excel_name):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_name
        workbook.save(excel_name)
    # 生成对应Excel
    workbook = openpyxl.load_workbook(excel_name)
    try:
        worksheet = workbook[sheet_name]
    except KeyError:
        worksheet = workbook.create_sheet(sheet_name)
    # 将行信息列表写入指定行
    row_num = 1  # 指定行号
    for col_num, value in enumerate(row_names, start=1):
        worksheet.cell(row=row_num, column=col_num, value=value)
    # 保存Excel文件
    workbook.save(excel_name)


def input_data(data, sheet_name):
    """
    打开对应sheet,输入data
    :param sheet_name:
    :param data:
    """
    workbook = openpyxl.load_workbook('compare_goals_actions.xlsx')
    if sheet_name == 'All Example':
        workbook = openpyxl.load_workbook('compare_goals_actions.xlsx')
        worksheet = workbook['All Example']

    else:
        worksheet = workbook.create_sheet(sheet_name)
        row_names = ['Actions', 'actions', 'Actions_number', 'actions_number', 'true_number', 'false_number']
        row_num = 1  # 指定行号
        for col_num, value in enumerate(row_names, start=1):
            worksheet.cell(row=row_num, column=col_num, value=value)

    # 保存Excel文件
    worksheet.append(data)
    workbook.save('compare_goals_actions.xlsx')


def write_data_to_excel(data, excel_name='compare_goals_actions.xlsx', sheet_name='All Example'):
    """
    按行输入数据
    data,
    excel_name = 'compare_goals_actions.xlsx',
    sheet_name = 'All Example'
    """
    if sheet_name == 'All Example':
        pass

    elif 'example' in sheet_name:
        pass


if __name__ == '__main__':
    write_row_name(['example', 'Instruction', 'Goals', 'Actions', 'answer', 'goals', 'actions', 'right_goal'])
    input_data()
    # print(example[0])
    # print(example[1])
